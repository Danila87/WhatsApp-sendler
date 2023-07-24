import tkinter as tk
import openpyxl
import re
import time
import Support_function
from prettytable import PrettyTable
from tkinter import filedialog
from colorama import Fore
from openpyxl.styles import PatternFill


class Excel:
    """ Класс для работы с Excel таблицей """

    def __init__(self):

        self.__wb = None
        self.__sheet_main = None
        self.__sheet_mailing_result = None

        self.__file_path = None

        self.__selected_columns = []
        self.__selected_columns_print = []

        self.select_excel()

    def __getattr__(self, item):
        return f'Атрибута {item} не существует'

    def __str__(self):
        return f'Выбранные колонки: {self.__selected_columns_print}\nТекущий файл: {self.__file_path}'

    def select_excel(self):

        """ Выбор файла Excel """

        self.__selected_columns = []
        self.__selected_columns_print = []

        root = tk.Tk()
        root.withdraw()

        self.__file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

        if not self.__file_path:
            print('Возникла ошибка! Возможно вы используете не тот файл!')

        self.__wb = openpyxl.load_workbook(self.__file_path)
        self.__sheet_main = self.__wb.active

    def select_column(self, clear: bool = False) -> list:

        """Выбор колонок с номерами"""

        if clear:
            self.__selected_columns = []
            self.__selected_columns_print = []

        if not self.__selected_columns:
            self.print_name_columns()

            while True:

                input_data = input(
                    'Введите номера колонок c номером телефона или введите "Далее" чтобы продолжить >> ').lower()

                if not input_data:
                    print(f'{Fore.RED}\nОшибка ввода данных!!\n')
                    continue

                if input_data in 'далее':
                    break

                input_data = re.findall(r'\d+', input_data)

                self.__selected_columns_print += [int(i) for i in input_data if 1 <= int(i) <= self.__sheet_main.max_column]
                self.__selected_columns_print = sorted(list(set(self.__selected_columns_print)))

                self.__selected_columns += [i - 1 for i in self.__selected_columns_print]
                self.__selected_columns = list(set(self.__selected_columns))

                print(f'\n{Fore.GREEN}Текущие колонки: {self.__selected_columns_print}')

        return self.__selected_columns

    def save_excel(self):

        """ Сохранение файла Excel """

        while True:

            try:
                self.__wb.save(self.__file_path)
                break

            except PermissionError:
                print(f'\n{Fore.RED}Не удалось сохранить файл. Закройте Excel таблицу, с данными.\n')

                for i in range(5, 0, -1):
                    print(f'Попытка повторного сохранения через {i} сек.')
                    time.sleep(1)

                print('\nПопытка сохранить файл......\n')
                time.sleep(2)

    def create_new_sheet(self):

        """ Создание дополнительного листа """

        if 'Mailing_result' not in self.__wb.sheetnames:
            self.__wb.create_sheet('Mailing_result')
            print('Страница результата создана.\n')

        else:
            self.__sheet_mailing_result = self.__wb['Mailing_result']
            self.__sheet_mailing_result.delete_rows(1, self.__sheet_mailing_result.max_row)
            print('Страница результата очищена и готова к записи.\n')

        self.save_excel()

    def write_additional_sheet(self, row: tuple, num_row: int, status: bool = True):

        """Запись результата рассылки на доп лист"""

        filling_successful = PatternFill(fill_type='solid', start_color='f5f5dc')
        filling_failed = PatternFill(fill_type='solid', start_color='c4c43f')

        self.__sheet_mailing_result.append(row)

        for cell in self.__sheet_mailing_result[num_row]:

            if not status:
                cell.fill = filling_failed
                continue

            cell.fill = filling_successful

        self.save_excel()

    def get_sheet_row(self) -> int:

        """Получить количество непустых строк"""

        num_data_rows = 0

        for i in self.__sheet_main.iter_rows(min_row=1, values_only=True):
            if any(i):
                num_data_rows += 1

        return num_data_rows

    def get_dict_keys_columns(self) -> dict:

        """Выбор колонок для ключей из Docx документа"""

        context_col = {}
        keys = Support_function.get_keys_from_word()

        if keys:
            self.print_name_columns()
            for i in keys:
                choice = input(f'Введите колонку для {i} >> ')
                context_col[i] = int(choice) - 1

        return context_col

    def print_name_columns(self):

        """Отображение превью колонок таблицы"""

        col_name = {}

        for column in self.__sheet_main.iter_rows(min_row=1, max_row=1, values_only=True):
            col_name = {i: col for i, col in enumerate(column) if col is not None}

        table = PrettyTable()
        table.field_names = [col + 1 for col in col_name.keys()]
        table.add_row([col for col in col_name.values()])

        print(table)

    @property
    def sheet_main(self):
        """
        Возврат главного листа
        """
        return self.__sheet_main

    @property
    def sheet_mailing_result(self):
        """
        Возврат листа с результатами рассылки
        """
        return self.__sheet_mailing_result


excel_table = Excel()
