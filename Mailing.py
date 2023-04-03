import openpyxl
import Excel_work as ex_work
import json
import requests

from colorama import init, Fore
from whatsapp_api_client_python import API
from openpyxl.styles import PatternFill


wb = openpyxl.load_workbook(filename='Table.xlsx')
sheet_start = wb['Start']

init(autoreset=True)

ID_INSTANCE = '1101799064'
GREEN_API_TOKEN = '13c3394261004255ad82faf458d8a36871ceaeedd0664cd3b8'

greenAPI = API.GreenApi(ID_INSTANCE, GREEN_API_TOKEN)


def main_mailing():

    """
    Главная функция рассылки, которая запускает вспомогательные функции.

    :return: Ничего не возвращает
    """

    column_length = ex_work.get_sheet_row()

    for row in range(2, column_length):
        keys = []
        for column in (7, 8):

            if sheet_start.cell(row=row, column=column).value is not None:

                name_organization = sheet_start.cell(row=row, column=4).value
                list_number = str(sheet_start.cell(row=row, column=column).value).split(';')
                inn = sheet_start.cell(row=row, column=3).value
                keys_send = []

                for number in list_number:

                    if check_number(name_organization=name_organization, phone_number=number) is True:
                        send_message(name_organization=name_organization, prone_number=number)
                    else:
                        # TODO Тут нужно доделать проверку, что если по всем номерам не получилось отправить смс,
                        #  то записывать этот момент

                        keys_send.append(False)
                        if len(keys) == 2 and all(keys):
                            print('Клиенту вообще не было отправлено смс')
                        unsent_messages(name_organization=name_organization, phone_number=number, inn=inn)
            else:
                # TODO доделать закрашивание всей строки где вообще нет номера в обеих колонках
                keys.append(False)
                if len(keys) == 2 and all(keys):
                    name_organization = sheet_start.cell(row=row, column=4).value
                    unsent_messages(name_organization=name_organization)
                    sheet_start.cell(row=row, column=1).fill = PatternFill(fill_type='solid', start_color='ffff00')
                    ex_work.save_excel()


def check_number(name_organization, phone_number) -> bool:

    """
    Функция проверяет Зарегистрирован ли номер в WhatsApp или нет

    :param name_organization: Название организации
    :param phone_number: Проверяемый номер телефона
    :return: Возвращает True или False в зависимости от результата. Если номер зарегистрирован то True,
     в обратном случае False
    """

    url = f"https://api.green-api.com/waInstance{ID_INSTANCE}/checkWhatsapp/{GREEN_API_TOKEN}"

    payload = json.dumps({"phoneNumber": phone_number})
    headers = {
        'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    if response.json()['existsWhatsapp'] is False:

        print(f'{Fore.RED}Организация {name_organization} по номеру {phone_number} не зарегистрирована в WhatsApp')
        return False

    elif response.json()['existsWhatsapp'] is True:

        return True


def send_message(name_organization: str, prone_number):

    """
    Функция рассылки сообщения

    :param name_organization: Название организации
    :param prone_number: Номер куда будет отправлено сообщение
    :return: Ничего не возвращает
    """

    print(f'{Fore.GREEN}Организации {name_organization} на номер {prone_number} отправили сообщение.')

# TODO Функцию надо переделать, чтобы она записывала данные на новый лист в книге Excel
def unsent_messages(name_organization, phone_number, inn):

    """
    Функция для записи на дополнительный лист экселя неудачных попыток отправки номера и незарегистрированных номеров в
    WhatsApp.

    :param name_organization: Название организации
    :param phone_number: Номер куда было отправлено сообщение
    :param inn: ИНН организации
    :return: Ничего не возвращает
    """

    txt_unset_numbers = 'unset_numbers.txt'
    with open(txt_unset_numbers, 'a', encoding='utf-8') as uns_numbers:
        uns_numbers.write(
            f'Организация {name_organization} ИНН {inn} по номеру {phone_number} не зарегистрирована в WhatsApp\n')

# TODO Тут идут просто черновые функции где я пробую создавать/очищать и записывать данные на новый лист.
#  В последствии этот набросок будет использоваться для записи ошибок/неотправленных смс на новый лист
def create_new_sheet():
    if 'Unset_numbers' in wb.sheetnames:
        ws1 = wb['Unset_numbers']
        ws1.delete_rows(2, ws1.max_row)
        print('Страница очищена и готова к записи.')
        wb.save("Table.xlsx")
    else:
        wb.create_sheet('Unset_numbers')
        wb.save("Table.xlsx")
        print('Страница создана.')


def write_to_excel(status: str):
    create_new_sheet()

    rows = ex_work.get_sheet_row()
    ws1 = wb['Unset_numbers']

    for i in range(1, rows):
        if sheet_start.cell(row=i, column=2).value == 'Бархатова Полина Павловна':
            ws1.append([sheet_start.cell(row=i, column=1).value, sheet_start.cell(row=i, column=2).value, sheet_start.cell(row=i, column=4).value, sheet_start.cell(row=i, column=3).value, status])
            wb.save("Table.xlsx")
        else:
            continue
